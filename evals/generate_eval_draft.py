"""Generate eval question draft from corpus courses → Excel for human review.

Samples questions only from courses listed in tamu_data/evals/eval_corpus.json.
Exports to tamu_data/evals/drafts/eval_draft_{YYYYMMDD}.xlsx for user review.

Usage:
    python evals/generate_eval_draft.py --n 60
    python evals/generate_eval_draft.py --n 10 --dry-run
"""

import argparse
import json
import random
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from evals.generate_golden_set import (
    CATEGORY_PROBS,
    CATEGORY_WEIGHTS,
    OUT_OF_SCOPE_QUESTIONS,
    STRATUM_MAP,
    synthesize_stratum,
)

CORPUS_PATH = Path("tamu_data/evals/eval_corpus.json")
DRAFTS_DIR = Path("tamu_data/evals/drafts")


def load_corpus() -> tuple[list[str], str]:
    """Load CRNs and department from eval_corpus.json."""
    if not CORPUS_PATH.exists():
        print(f"ERROR: Corpus file not found: {CORPUS_PATH}")
        print("Create it by editing tamu_data/evals/eval_corpus.json with ~25 CRNs.")
        sys.exit(1)
    with CORPUS_PATH.open() as f:
        data = json.load(f)
    crns = [str(c) for c in data.get("crns", [])]
    dept = data.get("department", "CSCE")
    if not crns:
        print("ERROR: eval_corpus.json has no CRNs. Add ~25 CRNs first.")
        sys.exit(1)
    return crns, dept


def sample_corpus_chunks(n_total: int, corpus_crns: list[str], department: str) -> list[dict]:
    """Sample chunks filtered to corpus CRNs only, weighted by CATEGORY_WEIGHTS."""
    from pymongo import MongoClient

    import config

    client = MongoClient(config.MONGODB_URI)
    db = client[config.MONGODB_DB]
    chunks_col = db["chunks"]  # v1 collection — used for category-weighted question synthesis only

    sampled: list[dict] = []
    for cat, prob in CATEGORY_PROBS.items():
        n_cat = max(2, round(prob * n_total))
        pipeline = [
            {
                "$match": {
                    "category": cat,
                    "crn": {"$in": corpus_crns},
                    "course_id": {"$regex": f"^{department}", "$options": "i"},
                    "content": {"$exists": True, "$ne": ""},
                }
            },
            {"$sample": {"size": n_cat}},
            {
                "$project": {
                    "_id": 0,
                    "crn": 1,
                    "course_id": 1,
                    "category": 1,
                    "title": 1,
                    "content": 1,
                    "section": 1,
                    "term": 1,
                    "instructor_name": 1,
                }
            },
        ]
        docs = list(chunks_col.aggregate(pipeline))
        for d in docs:
            d["_sampled_category"] = cat
        sampled.extend(docs)
        print(f"  {cat:<30} target={n_cat}  got={len(docs)}")

    client.close()
    print(f"  Total: {len(sampled)} chunks from {len(corpus_crns)} corpus CRNs")
    return sampled


def export_to_excel(questions: list[dict], output_path: Path) -> None:
    """Export questions to Excel with user-editable columns."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        print("ERROR: openpyxl required. Run: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Questions"

    # (header, width, user_editable)
    columns = [
        ("approved", 10, True),
        ("question", 60, True),
        ("reference_answer", 50, True),
        ("human_judgment", 15, True),
        ("stratum", 25, False),
        ("category", 25, False),
        ("source_crn", 12, False),
        ("source_course_id", 18, False),
        ("expected_function", 25, False),
    ]

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    readonly_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    wrap = Alignment(wrap_text=True, vertical="top")

    for col_idx, (name, width, _) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 20

    for row_idx, q in enumerate(questions, start=2):
        values = [
            True,  # approved
            q.get("question", ""),
            q.get("reference_answer", ""),
            "",  # human_judgment — blank, user fills 0/1 for RAGAS validation
            q.get("stratum", ""),
            q.get("category") or "",
            q.get("source_crn") or "",
            q.get("source_course_id") or "",
            q.get("expected_function", ""),
        ]
        for col_idx, (val, (_, _, editable)) in enumerate(zip(values, columns), start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = wrap
            if not editable:
                cell.fill = readonly_fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"  Saved: {output_path}")


def main():
    parser = argparse.ArgumentParser(
        description="Generate eval question draft from corpus courses → Excel"
    )
    parser.add_argument("--n", type=int, default=60, help="Target questions to generate (default: 60)")
    parser.add_argument("--dry-run", action="store_true",
                        help="Validate corpus + show planned distribution, skip synthesis")
    parser.add_argument("--seed", type=int, default=42)
    args = parser.parse_args()

    crns, dept = load_corpus()
    print(f"\nCorpus: {len(crns)} CRNs  |  Dept: {dept}  |  Target: {args.n} questions")

    if args.dry_run:
        print(f"\nDRY-RUN: Would generate {args.n} questions from {len(crns)} corpus CRNs.")
        print(f"CRNs: {crns}")
        base_total = sum(s["n_questions"] for s in STRATUM_MAP.values())
        scale = args.n / base_total
        print(f"\n  {'Stratum':<28} {'n':>4}")
        print(f"  {'-'*34}")
        for s, spec in STRATUM_MAP.items():
            n = max(1, round(spec["n_questions"] * scale))
            print(f"  {s:<28} {n:>4}")
        return

    rng = random.Random(args.seed)

    base_total = sum(s["n_questions"] for s in STRATUM_MAP.values())
    scale = args.n / base_total
    stratum_counts = {k: max(1, round(v["n_questions"] * scale)) for k, v in STRATUM_MAP.items()}

    n_retrieval = sum(v for k, v in stratum_counts.items() if k != "out_of_scope")
    n_chunks = n_retrieval * 4

    print(f"\n[1/3] Sampling ~{n_chunks} chunks from corpus CRNs...")
    all_chunks = sample_corpus_chunks(n_total=n_chunks, corpus_crns=crns, department=dept)

    print("\n[2/3] Synthesizing questions...")
    all_questions: list[dict] = []

    for stratum, spec in STRATUM_MAP.items():
        if stratum == "out_of_scope":
            continue
        n_q = stratum_counts[stratum]
        print(f"\n  [{stratum}]  n={n_q}  --  {spec['description']}")
        qs = synthesize_stratum(stratum, spec, all_chunks, n_q, rng)
        all_questions.extend(qs)

    # Add OOS
    n_oos = stratum_counts["out_of_scope"]
    oos_items = [
        {
            "question": q,
            "expected_function": "out_of_scope",
            "expected_course_ids": [],
            "expected_specific_categories": [],
            "expected_semantic_intent": False,
            "source_crn": None,
            "source_course_id": None,
            "source_category": None,
            "reference_answer": "(out of scope)",
            "stratum": "out_of_scope",
            "category": None,
        }
        for q in rng.sample(OUT_OF_SCOPE_QUESTIONS, min(n_oos, len(OUT_OF_SCOPE_QUESTIONS)))
    ]
    all_questions.extend(oos_items)
    print(f"\n  Added {n_oos} out_of_scope questions.")

    print(f"\n[3/3] Exporting {len(all_questions)} questions to Excel...")
    ts = datetime.now().strftime("%Y%m%d")
    out_path = DRAFTS_DIR / f"eval_draft_{ts}.xlsx"
    export_to_excel(all_questions, out_path)

    print(f"\nDone. Review and edit the draft:")
    print(f"  {out_path}")
    print(f"\n  - 'approved' = False to reject a question")
    print(f"  - Edit 'question' or 'reference_answer' to correct them")
    print(f"  - Fill 'human_judgment' (0=bad, 1=good) for RAGAS validation (optional)")
    print(f"\nThen import: python evals/import_eval_draft.py --draft {out_path} --tag v1")


if __name__ == "__main__":
    main()
