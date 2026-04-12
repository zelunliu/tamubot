"""Shared golden set I/O — Excel-native, no JSONL.

Schema columns (in order):
    id, question, reference_answer, expected_function, human_notes

run:<experiment> columns are appended by eval scripts and ignored by load().
"""
from __future__ import annotations

from pathlib import Path

SCHEMA_COLUMNS = ["id", "question", "reference_answer", "expected_function", "human_notes"]


def load(path: Path) -> list[dict]:
    """Read a golden set .xlsx. Returns list of dicts with SCHEMA_COLUMNS keys.

    run:<experiment> columns and any other non-schema columns are ignored.
    Rows with an empty question are skipped.
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl required: pip install openpyxl")

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        d = dict(zip(headers, row))
        question = str(d.get("question") or "").strip()
        if not question:
            continue
        items.append({
            "id": d.get("id"),
            "question": question,
            "reference_answer": str(d.get("reference_answer") or "").strip(),
            "expected_function": str(d.get("expected_function") or "").strip(),
            "human_notes": d.get("human_notes"),
        })
    return items


def save(items: list[dict], path: Path) -> None:
    """Write items to .xlsx with SCHEMA_COLUMNS column order.

    Creates parent directories if needed. Overwrites existing file.
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl required: pip install openpyxl")

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(SCHEMA_COLUMNS)
    for item in items:
        ws.append([item.get(col) for col in SCHEMA_COLUMNS])
    wb.save(path)


def append_run_column(path: Path, experiment: str, results: dict) -> None:
    """Add or overwrite column run:<experiment> in existing .xlsx.

    Matches rows by the id column. results maps question id -> value string.
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl required: pip install openpyxl")

    path = Path(path)
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    col_name = f"run:{experiment}"

    if col_name in headers:
        col_idx = headers.index(col_name) + 1  # openpyxl is 1-based
    else:
        col_idx = len(headers) + 1
        ws.cell(row=1, column=col_idx, value=col_name)

    id_col_idx = (headers.index("id") + 1) if "id" in headers else None

    for row_idx in range(2, ws.max_row + 1):
        if id_col_idx is None:
            break
        row_id = ws.cell(row=row_idx, column=id_col_idx).value
        if row_id in results:
            ws.cell(row=row_idx, column=col_idx, value=results[row_id])

    wb.save(path)
    wb.close()
