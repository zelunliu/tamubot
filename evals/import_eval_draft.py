"""Import approved Excel draft → Excel golden set.

Reads approved rows from a reviewed eval draft Excel and saves to
tamu_data/evals/golden_sets/golden_{YYYYMMDD}_{tag}.xlsx.

Usage:
    python evals/import_eval_draft.py --draft tamu_data/evals/drafts/eval_draft_20260311.xlsx --tag v1
    python evals/import_eval_draft.py --draft ... --tag v1 --allow-warnings
"""

import argparse
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from evals.golden_set import save as _save_golden_set

GOLDEN_SETS_DIR = Path("tamu_data/evals/golden_sets")
REQUIRED_FIELDS = ["question", "expected_function"]


def read_draft_excel(path: Path) -> list[dict]:
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl required. Run: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        rows.append(dict(zip(headers, row)))
    return rows


def coerce_bool(val) -> bool:
    """Coerce Excel bool/string/number to Python bool."""
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        return bool(val)
    if isinstance(val, str):
        return val.strip().lower() not in ("false", "0", "no", "")
    if val is None:
        return False
    return bool(val)


def validate_row(row: dict, row_num: int) -> list[str]:
    errors = []
    for field in REQUIRED_FIELDS:
        if not row.get(field):
            errors.append(f"Row {row_num}: missing '{field}'")
    question = str(row.get("question", "")).strip()
    if question and not question.endswith("?"):
        errors.append(f"Row {row_num}: question doesn't end with '?' — '{question[:50]}'")
    return errors


def main():
    parser = argparse.ArgumentParser(description="Import approved Excel draft → Excel golden set")
    parser.add_argument("--draft", required=True, help="Path to Excel draft file")
    parser.add_argument("--tag", default="v1", help="Tag suffix for output filename (default: v1)")
    parser.add_argument("--allow-warnings", action="store_true",
                        help="Import even if validation warnings exist")
    args = parser.parse_args()

    draft_path = Path(args.draft)
    if not draft_path.exists():
        print(f"ERROR: Draft file not found: {draft_path}")
        sys.exit(1)

    print(f"\nReading: {draft_path}")
    rows = read_draft_excel(draft_path)
    print(f"  Total rows: {len(rows)}")

    approved = [r for r in rows if coerce_bool(r.get("approved", True))]
    rejected = len(rows) - len(approved)
    print(f"  Approved: {len(approved)}  Rejected: {rejected}")

    if not approved:
        print("ERROR: No approved rows found.")
        sys.exit(1)

    # Validate
    all_errors = []
    for i, row in enumerate(approved, start=2):
        all_errors.extend(validate_row(row, i))

    if all_errors:
        for err in all_errors:
            print(f"  WARN: {err}")
        if not args.allow_warnings:
            print(f"\n{len(all_errors)} validation issue(s). Use --allow-warnings to import anyway.")
            sys.exit(1)

    # Build golden items
    golden_items = []
    for row in approved:
        item = {
            "id": row.get("id") or (len(golden_items) + 1),
            "question": str(row.get("question", "")).strip(),
            "reference_answer": str(row.get("reference_answer", "")).strip(),
            "expected_function": str(row.get("expected_function", "")).strip(),
            "human_notes": row.get("human_notes") or row.get("human_judgment"),
        }
        golden_items.append(item)

    # Save
    GOLDEN_SETS_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d")
    out_path = GOLDEN_SETS_DIR / f"golden_{ts}_{args.tag}.xlsx"
    _save_golden_set(golden_items, out_path)

    print(f"\nGolden set saved: {out_path}")
    print(f"  Items: {len(golden_items)}")

    print(f"\nNext: python evals/eval_chunking.py --golden-set {out_path} --experiment <name>")


if __name__ == "__main__":
    main()
