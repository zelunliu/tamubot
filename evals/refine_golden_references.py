"""Refine reference answers in a golden set JSONL using TAMU API synthesis.

For each entry, loads the full syllabus markdown for source_crn and calls the
TAMU API to synthesize a proper, concise answer that a course assistant should give.

Schema changes applied:
  - Removes outdated fields: stratum, category (redundant with source_category)
  - Keeps: question, reference_answer, source_crn, source_course_id, source_category,
           expected_function, expected_course_ids, expected_specific_categories,
           expected_semantic_intent, human_judgment

Usage:
    python evals/refine_golden_references.py
    python evals/refine_golden_references.py --input tamu_data/evals/golden_sets/golden_20260313_draft_v1.jsonl
    python evals/refine_golden_references.py --dry-run       # prints prompts, skips API
    python evals/refine_golden_references.py --overwrite     # writes in-place
    python evals/refine_golden_references.py --excel         # also export to .xlsx
    python evals/refine_golden_references.py --excel-only    # convert existing JSONL → .xlsx, no API calls
"""

import argparse
import json
import sys
import time
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import config

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

GOLDEN_DIR = Path("tamu_data/evals/golden_sets")
MARKDOWN_DIR = Path("tamu_data/processed/v3_step1_markdown")
DEFAULT_INPUT = GOLDEN_DIR / "golden_20260313_draft_v1.jsonl"
DEFAULT_OUTPUT = GOLDEN_DIR / "golden_20260410_v2.jsonl"

# ---------------------------------------------------------------------------
# Schema: fields to keep, in output order
# ---------------------------------------------------------------------------

KEEP_FIELDS = [
    "id",
    "question",
    "reference_answer",
    "source_crn",
    "source_course_id",
    "source_category",
    "expected_function",
    "expected_course_ids",
    "expected_specific_categories",
    "expected_semantic_intent",
    "human_judgment",
]

# Answers considered empty/unusable — checked as substrings of the lowercased answer.
# Keep patterns specific: broad phrases like "no information" match valid hedging language.
EMPTY_ANSWER_PATTERNS = [
    "(out of scope)",
    "does not contain any information",
    "no information about",
    "not available in the syllabus",
]

# ---------------------------------------------------------------------------
# Syllabus loader
# ---------------------------------------------------------------------------

def load_markdown_index() -> dict[str, str]:
    """Return {crn_str: full_markdown_text} for all v010 markdown files."""
    index: dict[str, str] = {}
    for md_path in MARKDOWN_DIR.glob("*.md"):
        # filename: 202611_CSCE_611_600_50668_v010.md  → CRN = 50668
        parts = md_path.stem.split("_")
        # last meaningful part before version tag is the CRN
        # pattern: TERM_DEPT_NUM_SECTION_CRN_vVER  (CRN is 5th field, 0-indexed)
        try:
            crn = parts[4]   # e.g. "50668"
        except IndexError:
            continue
        # Prefer v010 — skip if a v010 is already loaded
        version = parts[-1]  # e.g. "v010"
        if crn in index and version != "v010":
            continue
        text = md_path.read_text(encoding="utf-8", errors="replace")
        index[crn] = text
    return index


# ---------------------------------------------------------------------------
# Synthesis prompt builders
# ---------------------------------------------------------------------------

SYSTEM_PROMPT = (
    "You are writing reference answers for a TAMU course assistant evaluation suite. "
    "Each reference answer should be what an ideal, helpful assistant would say in response "
    "to a student's question — clear, accurate, and grounded strictly in the syllabus content "
    "provided. Do not add opinions, warnings, or information not present in the syllabus."
)


def _truncate(text: str, max_chars: int = 6000) -> str:
    if len(text) <= max_chars:
        return text
    return text[:max_chars] + "\n[...truncated]"


def build_prompt(item: dict, syllabus: str) -> str:
    """Build synthesis prompt based on expected_function and question type."""
    question = item["question"]
    course_id = item.get("source_course_id", "the course")
    expected_fn = item.get("expected_function", "")
    source_category = item.get("source_category", "")

    syllabus_block = _truncate(syllabus)

    if expected_fn == "out_of_scope":
        return ""  # handled separately

    if expected_fn == "recurrent":
        return f"""\
Student question: {question}

This question asks what courses pair with or follow {course_id}. The reference answer should
briefly describe what {course_id} covers (to establish context) and then note that the assistant
can help search for complementary courses — without naming specific courses that aren't mentioned
in the syllabus below.

Syllabus for {course_id}:
---
{syllabus_block}
---

Write a 2–4 sentence reference answer. Be factual. Do not fabricate course recommendations.
"""

    if expected_fn == "semantic_general":
        return f"""\
Student question: {question}

This is a cross-course discovery question. The reference answer should draw from the syllabus
content of {course_id} as a relevant example, and provide a factual answer that directly
addresses what the student asked.

Syllabus for {course_id} (relevant section: {source_category}):
---
{syllabus_block}
---

Write a concise, factual reference answer (2–5 sentences). Cite {course_id} as an example
where appropriate. Do not fabricate information about other courses not shown here.
"""

    # hybrid_course (default, advisory, specific, combined) — most entries
    advisory_note = ""
    if item.get("expected_semantic_intent"):
        advisory_note = (
            "The question is advisory or evaluative. Provide the relevant factual information "
            "from the syllabus that helps the student decide, without adding personal opinions. "
        )

    return f"""\
Student question: {question}

{advisory_note}Answer the question using only the syllabus content below for {course_id}.
The answer should be direct, factual, and complete — not a raw paste of syllabus text.
Write in a helpful assistant voice. If the specific information is not in the syllabus, say so briefly.

Syllabus for {course_id} (relevant section: {source_category}):
---
{syllabus_block}
---

Write the reference answer now. Be concise (under 150 words unless the question requires detail).
"""


# ---------------------------------------------------------------------------
# TAMU API call
# ---------------------------------------------------------------------------

def synthesize_answer(prompt: str, dry_run: bool = False) -> str:
    """Call TAMU API (streaming, as required by the gateway) and return the answer."""
    if dry_run:
        return "[DRY-RUN — no API call made]"

    tamu = config.get_tamu_client()
    stream = tamu.chat.completions.create(
        model=config.TAMU_MODEL,
        messages=[
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": prompt},
        ],
        temperature=0.2,
        max_tokens=4096,
        stream=True,
    )
    return "".join(chunk.choices[0].delta.content or "" for chunk in stream).strip()


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

# (column_name, width, user_editable)
EXCEL_COLUMNS: list[tuple[str, int, bool]] = [
    ("id",                           6, False),
    ("question",                    62, True),
    ("reference_answer",            55, True),
    ("human_judgment",              15, True),
    ("expected_function",           22, False),
    ("source_course_id",            18, False),
    ("source_category",             22, False),
    ("source_crn",                  12, False),
    ("expected_course_ids",         25, False),
    ("expected_specific_categories",28, False),
    ("expected_semantic_intent",    20, False),
]


def export_to_excel(items: list[dict], output_path: Path) -> None:
    """Export golden set items to a formatted .xlsx file."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        print("  [ERROR] openpyxl required: pip install openpyxl")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Golden Set"

    header_fill   = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    readonly_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    header_font   = Font(bold=True, color="FFFFFF")
    wrap          = Alignment(wrap_text=True, vertical="top")

    # Header row
    for col_idx, (name, width, _) in enumerate(EXCEL_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = wrap
        ws.column_dimensions[cell.column_letter].width = width
    ws.row_dimensions[1].height = 20

    # Data rows
    for row_idx, item in enumerate(items, start=2):
        for col_idx, (name, _, editable) in enumerate(EXCEL_COLUMNS, start=1):
            raw = item.get(name)
            # Render lists as comma-separated strings for readability
            if isinstance(raw, list):
                value = ", ".join(str(v) for v in raw) if raw else ""
            elif raw is None:
                value = ""
            else:
                value = raw
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = wrap
            if not editable:
                cell.fill = readonly_fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"  Excel: {output_path}")


# ---------------------------------------------------------------------------
# Schema cleanup + post-processing
# ---------------------------------------------------------------------------

def clean_item(item: dict, new_reference_answer: str) -> dict:
    """Return a new dict with only KEEP_FIELDS, updated reference_answer."""
    out = {k: item.get(k) for k in KEEP_FIELDS}
    out["reference_answer"] = new_reference_answer
    return out


def _is_empty_answer(answer: str) -> bool:
    """Return True if the answer is empty or contains a no-information pattern."""
    if not answer or not answer.strip():
        return True
    lower = answer.lower()
    return any(p in lower for p in EMPTY_ANSWER_PATTERNS)


def deduplicate_and_number(items: list[dict]) -> tuple[list[dict], int, int]:
    """Remove duplicate questions and empty answers, then assign sequential ids.

    Returns (cleaned_items, n_removed_dupes, n_removed_empty).
    """
    seen_questions: set[str] = set()
    n_dupes = 0
    n_empty = 0
    kept: list[dict] = []

    for item in items:
        q = item.get("question", "")
        ans = item.get("reference_answer", "")

        if q in seen_questions:
            n_dupes += 1
            continue
        if _is_empty_answer(ans):
            n_empty += 1
            continue

        seen_questions.add(q)
        kept.append(item)

    for i, item in enumerate(kept, start=1):
        item["id"] = i

    return kept, n_dupes, n_empty


# ---------------------------------------------------------------------------
# Main pipeline
# ---------------------------------------------------------------------------

def refine(
    input_path: Path,
    output_path: Path,
    dry_run: bool = False,
    verbose: bool = False,
    excel: bool = False,
    excel_only: bool = False,
) -> None:
    print(f"\n{'=' * 60}")
    print("  Golden Set Reference Answer Refinement")
    print(f"  Input:  {input_path}")
    print(f"  Output: {output_path}")
    if excel_only:
        print("  Mode:   Excel export only (no API calls)")
    else:
        print(f"  API:    {'DRY-RUN (no calls)' if dry_run else config.TAMU_MODEL}")
    print(f"{'=' * 60}\n")

    items = [json.loads(line) for line in input_path.read_text(encoding="utf-8").splitlines() if line.strip()]
    print(f"Loaded {len(items)} entries.")

    # --- Excel-only: clean + export, skip synthesis ---
    if excel_only:
        items, n_dupes, n_empty = deduplicate_and_number(items)
        print(f"  Removed — duplicates: {n_dupes}  |  empty answers: {n_empty}")
        print(f"  Remaining: {len(items)} entries")
        # Overwrite the JSONL with cleaned + numbered version
        output_path.parent.mkdir(parents=True, exist_ok=True)
        with output_path.open("w", encoding="utf-8") as f:
            for r in items:
                f.write(json.dumps(r, ensure_ascii=False) + "\n")
        print(f"  JSONL:  {output_path}")
        xlsx_path = output_path.with_suffix(".xlsx")
        export_to_excel(items, xlsx_path)
        return

    print(f"Indexing markdown files from {MARKDOWN_DIR}...")
    markdown_index = load_markdown_index()
    print(f"  {len(markdown_index)} syllabi indexed.\n")

    results = []
    n_synthesized = 0
    n_kept = 0
    n_no_markdown = 0

    for i, item in enumerate(items, 1):
        question = item.get("question", "")
        expected_fn = item.get("expected_function", "")
        crn = str(item.get("source_crn", "") or "")
        course_id = item.get("source_course_id", "")

        print(f"[{i:2d}/{len(items)}] {question[:70]}")
        print(f"         fn={expected_fn}  crn={crn}  cat={item.get('source_category', '')}")

        # --- out_of_scope: keep as-is ---
        if expected_fn == "out_of_scope":
            print("         → kept (out_of_scope)\n")
            results.append(clean_item(item, "(out of scope)"))
            n_kept += 1
            continue

        # --- no CRN: can't look up syllabus ---
        if not crn or crn == "None":
            print("         → kept (no source CRN)\n")
            results.append(clean_item(item, item.get("reference_answer", "")))
            n_kept += 1
            continue

        # --- find markdown ---
        syllabus = markdown_index.get(crn)
        if not syllabus:
            print(f"         [WARN] No markdown found for CRN {crn} — keeping original\n")
            results.append(clean_item(item, item.get("reference_answer", "")))
            n_no_markdown += 1
            continue

        # --- build prompt and synthesize ---
        prompt = build_prompt(item, syllabus)

        if verbose and dry_run:
            print(f"         --- PROMPT ---\n{prompt[:400]}\n         ---\n")

        try:
            answer = synthesize_answer(prompt, dry_run=dry_run)
            print(f"         → synthesized ({len(answer)} chars)\n")
            if verbose:
                print(f"         {answer[:200]}\n")
        except Exception as e:
            print(f"         [ERROR] Synthesis failed: {e} — keeping original\n")
            answer = item.get("reference_answer", "")

        results.append(clean_item(item, answer))
        n_synthesized += 1

        # Gentle rate-limiting between API calls
        if not dry_run and i < len(items):
            time.sleep(0.5)

    # --- Deduplicate, drop empty answers, assign ids ---
    results, n_dupes, n_empty = deduplicate_and_number(results)

    # --- Write output ---
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as f:
        for r in results:
            f.write(json.dumps(r, ensure_ascii=False) + "\n")

    print(f"{'=' * 60}")
    print(f"  Done: {len(results)} entries written to {output_path}")
    print(f"  Synthesized: {n_synthesized}  |  Kept: {n_kept}  |  No markdown: {n_no_markdown}")
    print(f"  Removed — duplicates: {n_dupes}  |  empty answers: {n_empty}")
    print(f"{'=' * 60}\n")

    if excel:
        xlsx_path = output_path.with_suffix(".xlsx")
        export_to_excel(results, xlsx_path)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Refine golden set reference answers via TAMU API synthesis"
    )
    parser.add_argument(
        "--input", default=str(DEFAULT_INPUT),
        help=f"Input JSONL path (default: {DEFAULT_INPUT})",
    )
    parser.add_argument(
        "--output", default=str(DEFAULT_OUTPUT),
        help=f"Output JSONL path (default: {DEFAULT_OUTPUT})",
    )
    parser.add_argument(
        "--overwrite", action="store_true",
        help="Write output to the same path as input (in-place refinement)",
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Build prompts and print stats without making API calls",
    )
    parser.add_argument(
        "--verbose", action="store_true",
        help="Print prompt previews and synthesized answers",
    )
    parser.add_argument(
        "--excel", action="store_true",
        help="Also export the output to a .xlsx file alongside the JSONL",
    )
    parser.add_argument(
        "--excel-only", action="store_true",
        help="Convert the input JSONL to .xlsx without making any API calls",
    )
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"[ERROR] Input not found: {input_path}")
        sys.exit(1)

    output_path = input_path if args.overwrite else Path(args.output)

    refine(
        input_path=input_path,
        output_path=output_path,
        dry_run=args.dry_run,
        verbose=args.verbose,
        excel=args.excel,
        excel_only=args.excel_only,
    )


if __name__ == "__main__":
    main()
