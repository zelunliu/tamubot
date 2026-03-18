"""End-to-end benchmark runner → versioned Excel + Markdown reports.

Runs a golden set JSONL through the full RAG pipeline, captures per-stage metrics,
and exports versioned reports for A/B experiment comparison.

Usage:
    python evals/run_benchmark.py \
        --golden-set tamu_data/evals/golden_sets/golden_20260311_v1.jsonl \
        --experiment-name cs600_ov100 \
        [--ragas]

Output:
    tamu_data/evals/reports/benchmark_{experiment_name}_{YYYYMMDD}.xlsx   (3 tabs)
    tamu_data/evals/reports/benchmark_{experiment_name}_{YYYYMMDD}.md
"""

import argparse
import json
import re
import subprocess
import sys
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Optional

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

import config
from rag import RouterResult, generator_order
from rag.v4.pipeline_v4 import run_pipeline_v4

REPORTS_DIR = Path("tamu_data/evals/reports")


# ---------------------------------------------------------------------------
# Per-question result
# ---------------------------------------------------------------------------

@dataclass
class BenchmarkRow:
    # Identity
    question_id: int
    question: str
    stratum: str
    source_course_id: str

    # Ground truth
    expected_function: str

    # Router — function + correctness
    router_function: str
    router_function_correct: bool

    # Router — extracted fields
    router_rewritten_query: str
    router_course_ids: str   # comma-separated, e.g. "CSCE 670, CSCE 638"
    router_intent_type: str  # empty string when None (out_of_scope)

    # Retrieval
    chunks_retrieved: int    # number of chunks returned after reranking

    # Token estimates (chars/4 — TAMU SSE doesn't expose counts)
    est_input_tokens: int    # query + all chunk content
    est_output_tokens: int   # answer length

    # Timing (ms)
    pipeline_ms: float       # router + retrieval combined (v4 graph)
    generator_ms: float
    total_ms: float

    # Generator
    answer_full: str
    answer_preview: str      # first 120 chars — see Full Answers tab for complete text
    citation_pass: bool

    # Per-node timing from pipeline timing_ms (inside-graph, not wall-clock)
    is_recurrent: bool = False
    router_ms: Optional[float] = None
    retrieval_ms: Optional[float] = None
    generator_node_ms: Optional[float] = None
    # Recurrent path only (None for non-recurrent rows)
    anchor_ms: Optional[float] = None
    eval_search_ms: Optional[float] = None
    schedule_filter_ms: Optional[float] = None
    merge_ms: Optional[float] = None

    # RAGAS (optional — populated with --ragas flag)
    ragas_faithfulness: Optional[float] = None
    ragas_relevancy: Optional[float] = None

    # Error
    error: Optional[str] = None


# ---------------------------------------------------------------------------
# Pipeline runner
# ---------------------------------------------------------------------------

def run_one(item: dict, do_ragas: bool, question_id: int = 0) -> BenchmarkRow:
    """Run one golden set item through the full v4 pipeline (router → retrieval → generation)."""
    query = item["question"]
    source_course_id = str(item.get("source_course_id") or "")
    expected_fn = str(item.get("expected_function", ""))

    t0 = time.perf_counter()
    error: Optional[str] = None
    chunks: list[dict] = []
    answer = ""
    rr = RouterResult(rewritten_query=query, category_confidence=0.0)
    data_gaps: list = []
    data_integrity = True
    conflicted_ids: list = []

    # Router + Retrieval (v4 graph)
    timing_ms: dict = {}
    try:
        chunks, rr_result, data_gaps, data_integrity, conflicted_ids, timing_ms = run_pipeline_v4(
            query, return_timing=True
        )
        if rr_result is not None:
            rr = rr_result
        else:
            error = "retrieval: router_result is None (pipeline did not complete)"
    except Exception as e:
        error = f"retrieval: {e}"
    pipeline_ms = round((time.perf_counter() - t0) * 1000, 1)

    # Per-node timing extraction
    is_recurrent = (rr.function == "recurrent")
    router_ms = timing_ms.get("router_node")
    retrieval_ms = timing_ms.get("retrieval_node")
    generator_node_ms = timing_ms.get("generator_node")
    anchor_ms = timing_ms.get("anchor_node") if is_recurrent else None
    eval_search_ms = timing_ms.get("eval_search_node") if is_recurrent else None
    schedule_filter_ms = timing_ms.get("schedule_filter_node") if is_recurrent else None
    merge_ms = timing_ms.get("merge_node") if is_recurrent else None

    # Generation
    t_gen = time.perf_counter()
    if not error:
        try:
            stream = generator_order(
                recurrent=is_recurrent,
                chunks=chunks,
                query=query,
                router_result=rr,
                data_gaps=data_gaps,
                data_integrity=data_integrity,
                conflicted_course_ids=conflicted_ids,
            )
            answer = "".join(stream)
        except Exception as e:
            error = f"generation: {e}"
    generator_ms = round((time.perf_counter() - t_gen) * 1000, 1)
    total_ms = round((time.perf_counter() - t0) * 1000, 1)

    # RAGAS (optional)
    ragas_faithfulness: Optional[float] = None
    ragas_relevancy: Optional[float] = None
    if do_ragas and chunks and answer and not error:
        try:
            from rag import compute_ragas_metrics
            contexts = [c.get("content", "") for c in chunks if c.get("content")]
            scores = compute_ragas_metrics(question=query, contexts=contexts, answer=answer)
            ragas_faithfulness = scores.get("faithfulness")
            ragas_relevancy = scores.get("answer_relevancy")
        except Exception:
            pass  # non-fatal

    return BenchmarkRow(
        question_id=question_id,
        question=query,
        stratum=item.get("stratum", ""),
        source_course_id=source_course_id,
        expected_function=expected_fn,
        router_function=rr.function,
        router_function_correct=(rr.function == expected_fn),
        router_rewritten_query=rr.rewritten_query or "",
        router_course_ids=", ".join(rr.course_ids) if rr.course_ids else "",
        router_intent_type=rr.intent_type or "",
        chunks_retrieved=len(chunks),
        est_input_tokens=max(0, (len(query) + sum(len(c.get("content", "")) for c in chunks)) // 4),
        est_output_tokens=len(answer) // 4,
        pipeline_ms=pipeline_ms,
        generator_ms=generator_ms,
        total_ms=total_ms,
        answer_full=answer,
        answer_preview=(answer[:120] + "…" if len(answer) > 120 else answer),
        citation_pass=bool(re.search(r"\[Source \d+\]", answer)),
        is_recurrent=is_recurrent,
        router_ms=router_ms,
        retrieval_ms=retrieval_ms,
        generator_node_ms=generator_node_ms,
        anchor_ms=anchor_ms,
        eval_search_ms=eval_search_ms,
        schedule_filter_ms=schedule_filter_ms,
        merge_ms=merge_ms,
        ragas_faithfulness=ragas_faithfulness,
        ragas_relevancy=ragas_relevancy,
        error=error,
    )


# ---------------------------------------------------------------------------
# Output: Excel (3 tabs) + Markdown
# ---------------------------------------------------------------------------

def _get_git_commit() -> str:
    try:
        return subprocess.check_output(
            ["git", "rev-parse", "--short", "HEAD"], stderr=subprocess.DEVNULL
        ).decode().strip()
    except Exception:
        return "unknown"


def _avg(rows: list[BenchmarkRow], attr: str) -> Optional[float]:
    vals = [getattr(r, attr) for r in rows if getattr(r, attr) is not None]
    return round(sum(vals) / len(vals), 1) if vals else None


def write_excel(
    rows: list[BenchmarkRow],
    experiment_name: str,
    golden_set_path: str,
    output_path: Path,
    do_ragas: bool,
) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        print("ERROR: openpyxl required. Run: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.Workbook()
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    wrap = Alignment(wrap_text=True, vertical="top")

    n = len(rows)
    n_fn_correct = sum(1 for r in rows if r.router_function_correct)
    citation_cases = [r for r in rows if r.router_function != "out_of_scope"]
    citation_pass = sum(1 for r in citation_cases if r.citation_pass)
    ragas_cases = [r for r in rows if r.ragas_faithfulness is not None]
    avg_faith = sum(r.ragas_faithfulness for r in ragas_cases) / len(ragas_cases) if ragas_cases else None

    # ── Tab 1: Summary ───────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"

    avg_relevancy = sum(r.ragas_relevancy for r in ragas_cases) / len(ragas_cases) if ragas_cases else None
    avg_chunks = _avg(rows, "chunks_retrieved")

    summary_rows = [
        ("Experiment", experiment_name),
        ("Date", datetime.now().strftime("%Y-%m-%d")),
        ("n_questions", n),
        ("Router accuracy",
         f"{n_fn_correct/n:.1%} ({n_fn_correct}/{n})" if n else "N/A"),
        ("Citation pass rate",
         f"{citation_pass/len(citation_cases):.1%} ({citation_pass}/{len(citation_cases)})" if citation_cases else "N/A"),
        ("Mean RAGAS faithfulness",
         f"{avg_faith:.2f}" if avg_faith is not None else "not run (use --ragas)"),
        ("Mean RAGAS relevancy",
         f"{avg_relevancy:.2f}" if avg_relevancy is not None else "not run (use --ragas)"),
        ("Mean chunks retrieved", f"{avg_chunks:.1f}" if avg_chunks is not None else "N/A"),
        ("Mean est. input tokens", _avg(rows, "est_input_tokens")),
        ("Mean est. output tokens", _avg(rows, "est_output_tokens")),
        ("Mean total latency (ms)", _avg(rows, "total_ms")),
        ("Mean pipeline latency (ms)", _avg(rows, "pipeline_ms")),
        ("Mean generator latency (ms)", _avg(rows, "generator_ms")),
        ("Errors", sum(1 for r in rows if r.error)),
    ]

    for col, label in [(1, "Metric"), (2, "Value")]:
        cell = ws_sum.cell(row=1, column=col, value=label)
        cell.font = header_font
        cell.fill = header_fill
    ws_sum.column_dimensions["A"].width = 35
    ws_sum.column_dimensions["B"].width = 30

    for i, (metric, val) in enumerate(summary_rows, start=2):
        ws_sum.cell(row=i, column=1, value=metric)
        ws_sum.cell(row=i, column=2, value=val)

    # ── Tab 2: Per-Query ─────────────────────────────────────────────────
    ws_pq = wb.create_sheet("Per-Query")

    pq_cols = [
        # Identity
        ("q#", 5),
        ("question", 50),
        ("stratum", 22),
        ("source_course_id", 18),
        # Routing ground truth + correctness
        ("expected_function", 20),
        ("router_function", 20),
        ("router_function_correct", 10),
        # Router extracted fields
        ("router_rewritten_query", 42),
        ("router_course_ids", 22),
        ("router_intent_type", 16),
        # Retrieval + tokens
        ("chunks_retrieved", 10),
        ("est_input_tokens", 14),
        ("est_output_tokens", 14),
        # Generation
        ("citation_pass", 10),
        ("ragas_faithfulness", 14),
        ("ragas_relevancy", 14),
        # Timing
        ("pipeline_ms", 12),
        ("generator_ms", 12),
        ("total_ms", 10),
        # Answer preview (short) + diagnostics
        ("answer_preview", 45),
        ("error", 30),
        # human_judgment column: blank, user fills in Full Answers tab
        ("human_judgment", 14),
    ]

    for col_idx, (h, width) in enumerate(pq_cols, start=1):
        cell = ws_pq.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        ws_pq.column_dimensions[cell.column_letter].width = width

    for row_idx, r in enumerate(rows, start=2):
        values = [
            r.question_id, r.question, r.stratum, r.source_course_id,
            r.expected_function, r.router_function, r.router_function_correct,
            r.router_rewritten_query, r.router_course_ids, r.router_intent_type,
            r.chunks_retrieved, r.est_input_tokens, r.est_output_tokens,
            r.citation_pass, r.ragas_faithfulness, r.ragas_relevancy,
            r.pipeline_ms, r.generator_ms, r.total_ms,
            r.answer_preview, r.error,
            "",  # human_judgment — blank for user to fill
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws_pq.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = wrap

    ws_pq.freeze_panes = "A2"
    ws_pq.auto_filter.ref = ws_pq.dimensions

    # ── Tab 3: Config ────────────────────────────────────────────────────
    ws_cfg = wb.create_sheet("Config")

    gen_model = (
        getattr(config, "TAMU_MODEL", "?")
        if getattr(config, "USE_TAMU_API", False)
        else getattr(config, "GENERATION_MODEL", "?")
    )
    cfg_rows = [
        ("experiment_name", experiment_name),
        ("date", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("git_commit", _get_git_commit()),
        ("golden_set_path", golden_set_path),
        ("n_questions", n),
        ("ragas_enabled", do_ragas),
        ("generation_model", gen_model),
        ("use_tamu_api", getattr(config, "USE_TAMU_API", False)),
    ]

    for col, label in [(1, "Key"), (2, "Value")]:
        cell = ws_cfg.cell(row=1, column=col, value=label)
        cell.font = header_font
        cell.fill = header_fill
    ws_cfg.column_dimensions["A"].width = 28
    ws_cfg.column_dimensions["B"].width = 50

    for i, (k, v) in enumerate(cfg_rows, start=2):
        ws_cfg.cell(row=i, column=1, value=k)
        ws_cfg.cell(row=i, column=2, value=str(v))

    # ── Tab 4: Full Answers ───────────────────────────────────────────────
    ws_fa = wb.create_sheet("Full Answers")

    fa_cols = [
        ("q#", 5),
        ("question", 50),
        ("answer_full", 110),
        ("ragas_faithfulness", 16),
        ("ragas_relevancy", 14),
        ("human_judgment", 14),
        ("error", 30),
    ]
    for col_idx, (label, width) in enumerate(fa_cols, start=1):
        cell = ws_fa.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        ws_fa.column_dimensions[cell.column_letter].width = width

    for row_idx, r in enumerate(rows, start=2):
        fa_values = [r.question_id, r.question, r.answer_full or "", r.ragas_faithfulness, r.ragas_relevancy, "", r.error]
        for col_idx, val in enumerate(fa_values, start=1):
            cell = ws_fa.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws_fa.row_dimensions[row_idx].height = 80  # fixed height — double-click row border to auto-fit

    ws_fa.freeze_panes = "A2"
    ws_fa.row_dimensions[1].height = 18

    # ── Tab 5: Column Definitions ─────────────────────────────────────────
    ws_defs = wb.create_sheet("Column Definitions")

    col_defs = [
        ("Column", "Type", "Description"),
        # Identity
        ("question", "str", "The question posed to the bot, from the golden set."),
        ("stratum", "str", "Sampling stratum used during golden set generation (e.g. metadata_default, semantic_general). Groups questions by query pattern for stratified accuracy analysis."),
        ("source_course_id", "str", "Course identifier of the golden source (e.g. 'CSCE 672'). Provenance reference."),
        # Ground truth
        ("expected_function", "str", "The router function the question should trigger (e.g. hybrid_course, semantic_general, out_of_scope). Ground truth label from the golden set."),
        # Router — function + correctness
        ("router_function", "str", "The routing function selected by the v4 pipeline. Values: hybrid_course (course-specific query), recurrent (cross-corpus discovery), semantic_general (broad search), out_of_scope (unrelated query)."),
        ("router_function_correct", "bool", "True if router_function == expected_function. Primary routing accuracy signal."),
        # Router — extracted fields (LLM output)
        ("router_rewritten_query", "str", "Query rewritten by the router LLM for better retrieval alignment. This is what gets embedded and sent to Voyage/MongoDB."),
        ("router_course_ids", "str", "Comma-separated course IDs extracted from the query (e.g. 'CSCE 670, CSCE 638'). Empty for general/discovery queries."),
        ("router_intent_type", "str", "Advisory/evaluative intent detected by the LLM. Empty = purely factual or out_of_scope. Values: ACADEMIC, CAREER, DIFFICULTY, PLANNING, ADMINISTRATIVE, GENERAL."),
        # Timing
        ("pipeline_ms", "float (ms)", "Wall-clock time for the v4 graph (router + retrieval combined)."),
        ("generator_ms", "float (ms)", "Wall-clock time for generation: context assembly + LLM streaming."),
        ("total_ms", "float (ms)", "Total wall-clock time from query receipt to answer complete."),
        # Generator
        ("citation_pass", "bool", "True if the generated answer contains at least one [Source N] citation. Regex check — confirms the generator cited something, not that it cited correctly."),
        ("ragas_faithfulness", "float 0–1 / blank",
         "RAGAS Faithfulness — measures whether every claim in the answer is grounded in the retrieved chunks.\n\n"
         "HOW IT WORKS (3 LLM calls):\n"
         "  1. Decompose: critic LLM reads the answer and extracts a list of atomic claims "
         "(e.g. 'The final exam is worth 40%', 'Late work is not accepted').\n"
         "  2. Verify: for each claim, the critic LLM is shown the retrieved chunks and asked "
         "whether the claim is fully supported, partially supported, or not supported by the context.\n"
         "  3. Score: faithfulness = supported_claims / total_claims.\n\n"
         "CRITIC MODEL: TAMU gateway (same model as the generator). Temperature=0.\n"
         "KNOWN LIMITATION: The critic is the same model family that wrote the answer, so it may be "
         "lenient on its own phrasing. This is mitigated by temperature=0 and structured prompts.\n\n"
         "THRESHOLDS: 0.95+ excellent | 0.85–0.95 good | 0.70–0.85 acceptable | <0.70 hallucination risk.\n"
         "Only populated when --ragas flag is used (~30s per question)."),
        ("ragas_relevancy", "float 0–1 / blank",
         "RAGAS Answer Relevancy — measures whether the answer actually addresses the question asked "
         "(vs. being technically faithful but off-topic).\n\n"
         "HOW IT WORKS (1 LLM call + embedding similarity):\n"
         "  1. Generate questions: critic LLM reads only the answer (not the original question) and "
         "generates N=3 hypothetical questions that this answer appears to be responding to.\n"
         "  2. Embed: both the N generated questions and the original question are embedded using Voyage-3.\n"
         "  3. Score: mean cosine similarity between the original question embedding and each generated "
         "question embedding. High similarity = the answer is specifically about what was asked.\n\n"
         "CRITIC MODEL: TAMU gateway LLM (question generation) + Voyage-3 (embeddings). "
         "Temperature=0.\n"
         "WHY IT'S NOISIER THAN FAITHFULNESS: similarity of question phrasings is inherently fuzzy. "
         "A correct but verbose answer may score lower than a concise one. "
         "An answer that says 'I don't know' scores near 0 because the generated questions won't match.\n\n"
         "THRESHOLDS: 0.85+ excellent | 0.75–0.85 good | 0.65–0.75 acceptable | <0.65 answer is drifting off-topic.\n"
         "Treat as directional — use alongside faithfulness and human_judgment, not as a standalone gate.\n"
         "Only populated when --ragas flag is used (~30s per question)."),
        ("router_ms + retrieval_ms + generator_ms", "—", "The three stages sum to approximately total_ms. Small gap = overhead from Python timing and RAGAS (if enabled)."),
        ("generator_ms", "float (ms)", "Wall-clock time for the generation LLM call: context assembly + LLM streaming + citation gate."),
        ("total_ms", "float (ms)", "Total wall-clock time from query receipt to answer complete (router + retrieval + generation)."),
        ("answer_preview", "str", "First 300 characters of the generated answer. See 'Full Answers' tab for complete text with RAGAS scores alongside each answer."),
        ("error", "str / blank", "Error message if any pipeline stage failed. Blank on success."),
        ("human_judgment", "blank → user fills", "Leave blank to fill in after reviewing the Full Answers tab. Suggested values: correct / partial / wrong / hallucinated."),
    ]

    def_widths = [28, 18, 100]
    for col_idx, (label, width) in enumerate(zip(["Column", "Type", "Description"], def_widths), start=1):
        cell = ws_defs.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        ws_defs.column_dimensions[cell.column_letter].width = width

    for row_idx, row_vals in enumerate(col_defs[1:], start=2):
        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws_defs.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws_defs.freeze_panes = "A2"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"  Excel: {output_path}")


def write_markdown(
    rows: list[BenchmarkRow],
    experiment_name: str,
    output_path: Path,
) -> None:
    n = len(rows)
    n_fn_correct = sum(1 for r in rows if r.router_function_correct)
    citation_cases = [r for r in rows if r.router_function != "out_of_scope"]
    citation_pass = sum(1 for r in citation_cases if r.citation_pass)
    ragas_cases = [r for r in rows if r.ragas_faithfulness is not None]
    avg_faith = sum(r.ragas_faithfulness for r in ragas_cases) / len(ragas_cases) if ragas_cases else None

    def _fmt_ms(attr: str) -> str:
        v = _avg(rows, attr)
        return f"{v:.0f}" if v is not None else "N/A"

    lines = [
        f"# Benchmark Report: {experiment_name}",
        "",
        f"**Date:** {datetime.now().strftime('%Y-%m-%d %H:%M')}  ",
        f"**Git commit:** {_get_git_commit()}  ",
        f"**Questions:** {n}",
        "",
        "## Summary",
        "",
        "| Metric | Value |",
        "|--------|-------|",
    ]

    avg_relevancy = sum(r.ragas_relevancy for r in ragas_cases) / len(ragas_cases) if ragas_cases else None

    if n:
        lines += [
            f"| Router accuracy | {n_fn_correct/n:.1%} ({n_fn_correct}/{n}) |",
            f"| Citation pass rate | {citation_pass/len(citation_cases):.1%} ({citation_pass}/{len(citation_cases)}) |"
            if citation_cases else "| Citation pass rate | N/A |",
            f"| Mean RAGAS faithfulness | {avg_faith:.2f} |"
            if avg_faith is not None else "| Mean RAGAS faithfulness | not run |",
            f"| Mean RAGAS relevancy | {avg_relevancy:.2f} |"
            if avg_relevancy is not None else "| Mean RAGAS relevancy | not run |",
            f"| Mean chunks retrieved | {_fmt_ms('chunks_retrieved')} |",
            f"| Mean est. input tokens | {_fmt_ms('est_input_tokens')} |",
            f"| Mean est. output tokens | {_fmt_ms('est_output_tokens')} |",
            f"| Mean total latency (ms) | {_fmt_ms('total_ms')} |",
            f"| Mean pipeline latency (ms) | {_fmt_ms('pipeline_ms')} |",
            f"| Mean generator latency (ms) | {_fmt_ms('generator_ms')} |",
            f"| Errors | {sum(1 for r in rows if r.error)} |",
        ]

    # Per-stratum breakdown
    lines += ["", "## Router Accuracy by Stratum", "",
              "| Stratum | Correct | Total | Accuracy |",
              "|---------|---------|-------|----------|"]
    strata = sorted({r.stratum for r in rows})
    for s in strata:
        s_rows = [r for r in rows if r.stratum == s]
        s_correct = sum(1 for r in s_rows if r.router_function_correct)
        lines.append(f"| {s} | {s_correct} | {len(s_rows)} | {s_correct/len(s_rows):.0%} |")

    # Errors
    errors = [r for r in rows if r.error]
    if errors:
        lines += ["", f"## Errors ({len(errors)})", ""]
        for r in errors:
            q = r.question[:70] + "..." if len(r.question) > 70 else r.question
            lines.append(f"- **{q}** → `{r.error}`")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text("\n".join(lines), encoding="utf-8")
    print(f"  Markdown: {output_path}")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Benchmark runner — runs golden set through pipeline, exports versioned reports"
    )
    parser.add_argument("--golden-set", required=True, help="Path to golden set JSONL")
    parser.add_argument("--experiment-name", required=True,
                        help="Experiment identifier embedded in output filename (e.g. cs600_ov100)")
    parser.add_argument("--ragas", action="store_true",
                        help="Run RAGAS faithfulness/relevancy scores (~30s per question)")
    args = parser.parse_args()

    golden_path = Path(args.golden_set)
    if not golden_path.exists():
        print(f"ERROR: Golden set not found: {golden_path}")
        sys.exit(1)

    items = []
    with golden_path.open(encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                items.append(json.loads(line))

    print(f"\nBenchmark: {args.experiment_name}")
    print(f"Golden set: {golden_path}  ({len(items)} items)")
    if args.ragas:
        print("RAGAS: enabled")

    rows: list[BenchmarkRow] = []
    for i, item in enumerate(items, start=1):
        if i > 1:
            time.sleep(2)  # avoid RPM spikes (proxy limit: 30/min)
        q_preview = item.get("question", "")[:60]
        print(f"\n[{i:3d}/{len(items)}] {q_preview}...", flush=True)
        row = run_one(item, do_ragas=args.ragas, question_id=i)
        rows.append(row)
        fn_ok = "v" if row.router_function_correct else "x"
        status = row.error or "OK"
        print(
            f"         fn={fn_ok} pipeline={row.pipeline_ms:.0f}ms "
            f"gen={row.generator_ms:.0f}ms  [{status}]"
        )

    # Write reports
    ts = datetime.now().strftime("%Y%m%d")
    base_name = f"benchmark_{args.experiment_name}_{ts}"
    xlsx_path = REPORTS_DIR / f"{base_name}.xlsx"
    md_path = REPORTS_DIR / f"{base_name}.md"

    print("\nWriting reports...")
    write_excel(rows, args.experiment_name, str(golden_path), xlsx_path, args.ragas)
    write_markdown(rows, args.experiment_name, md_path)

    # Final summary
    n = len(rows)
    n_correct = sum(1 for r in rows if r.router_function_correct)
    print(f"\n{'='*55}")
    print(f"  Experiment:      {args.experiment_name}")
    print(f"  Questions:       {n}")
    if n:
        print(f"  Router accuracy: {n_correct/n:.1%} ({n_correct}/{n})")
    errors = sum(1 for r in rows if r.error)
    if errors:
        print(f"  Errors:          {errors}")
    print(f"{'='*55}")
    print(f"\nTo validate RAGAS: python evals/validate_ragas.py --benchmark {xlsx_path}")


if __name__ == "__main__":
    main()
