"""Validate RAGAS scores against human judgments.

Reads the Per-Query tab of a benchmark Excel, correlates RAGAS metrics against
human_judgment column (0=bad, 1=good), and reports trustworthiness of each metric.

Usage:
    python evals/validate_ragas.py \
        --benchmark tamu_data/evals/reports/benchmark_cs600_ov100_20260311.xlsx \
        [--min-human-labels 20]

Workflow:
    1. Run: make bench-ragas GOLDEN=... EXP=...
    2. Open benchmark Excel → Per-Query tab → fill 'human_judgment' column (0/1)
    3. Run this script to compute correlation between RAGAS scores and human labels
"""

import argparse
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

REPORTS_DIR = Path("tamu_data/evals/reports")


def read_per_query_tab(path: Path) -> list[dict]:
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl required. Run: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(path)
    if "Per-Query" not in wb.sheetnames:
        print("ERROR: No 'Per-Query' tab found. Make sure this is a benchmark Excel file.")
        sys.exit(1)

    ws = wb["Per-Query"]
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        rows.append(dict(zip(headers, row)))
    return rows


def compute_correlations(x: list[float], y: list[float]) -> dict:
    try:
        from scipy import stats
    except ImportError:
        print("ERROR: scipy required. Run: pip install scipy")
        sys.exit(1)

    if len(x) < 3:
        return {"pearson_r": None, "pearson_p": None, "spearman_r": None, "spearman_p": None}

    pearson_r, pearson_p = stats.pearsonr(x, y)
    spearman_r, spearman_p = stats.spearmanr(x, y)
    return {
        "pearson_r": round(float(pearson_r), 3),
        "pearson_p": round(float(pearson_p), 4),
        "spearman_r": round(float(spearman_r), 3),
        "spearman_p": round(float(spearman_p), 4),
    }


def interpret(r: Optional[float]) -> str:
    if r is None:
        return "N/A"
    a = abs(r)
    if a > 0.6:
        return "strong"
    if a > 0.4:
        return "moderate"
    return "weak"


def main():
    parser = argparse.ArgumentParser(
        description="Validate RAGAS scores vs human judgments (Pearson + Spearman)"
    )
    parser.add_argument("--benchmark", required=True, help="Path to benchmark Excel file")
    parser.add_argument("--min-human-labels", type=int, default=20,
                        help="Min labeled rows required (default: 20)")
    args = parser.parse_args()

    bench_path = Path(args.benchmark)
    if not bench_path.exists():
        print(f"ERROR: Benchmark not found: {bench_path}")
        sys.exit(1)

    print(f"\nReading: {bench_path}")
    rows = read_per_query_tab(bench_path)
    print(f"  Total rows: {len(rows)}")

    # Filter to rows with human_judgment filled
    labeled = [
        r for r in rows
        if r.get("human_judgment") is not None and str(r.get("human_judgment", "")).strip() != ""
    ]
    print(f"  Rows with human_judgment: {len(labeled)}")

    if len(labeled) < args.min_human_labels:
        print(f"\nERROR: Need at least {args.min_human_labels} human-labeled rows, got {len(labeled)}.")
        print("\nHow to add labels:")
        print(f"  1. Open: {bench_path}")
        print("  2. Go to Per-Query tab")
        print("  3. Fill the 'human_judgment' column: 1=good answer, 0=bad answer")
        print("  4. Re-run this script")
        sys.exit(1)

    try:
        human = [float(r["human_judgment"]) for r in labeled]
    except (TypeError, ValueError) as e:
        print(f"ERROR: human_judgment values must be 0 or 1, got non-numeric: {e}")
        sys.exit(1)

    metrics_to_check = ["ragas_faithfulness", "ragas_relevancy"]

    results: dict[str, dict] = {}
    for metric in metrics_to_check:
        valid_pairs = [
            (h, float(r[metric]))
            for h, r in zip(human, labeled)
            if r.get(metric) is not None
        ]
        if len(valid_pairs) < 5:
            print(f"  SKIP {metric}: only {len(valid_pairs)} non-null values (need ≥5)")
            continue
        h_vals = [h for h, _ in valid_pairs]
        m_vals = [v for _, v in valid_pairs]
        corr = compute_correlations(h_vals, m_vals)
        corr["n"] = len(valid_pairs)
        results[metric] = corr

    if not results:
        print("\nNo RAGAS metrics available to validate.")
        print("Run: make bench-ragas GOLDEN=... EXP=... first.")
        sys.exit(1)

    # Print table
    print(f"\n{'='*72}")
    print("  RAGAS vs Human Judgment Correlation")
    print(f"  Benchmark: {bench_path.name}  |  Labeled: {len(labeled)} rows")
    print(f"{'='*72}")
    print(f"  {'Metric':<25} {'Pearson r':>10} {'p-value':>9} {'Spearman r':>12} {'p-value':>9} {'n':>5} {'Strength':>10}")
    print(f"  {'-'*72}")

    for metric, corr in results.items():
        sig_p = "✓" if (corr["pearson_p"] or 1) < 0.05 else "✗"
        strength = interpret(corr.get("pearson_r"))
        print(
            f"  {metric:<25} "
            f"{corr['pearson_r']:>10.3f} "
            f"{corr['pearson_p']:>8.4f}{sig_p} "
            f"{corr['spearman_r']:>12.3f} "
            f"{corr['spearman_p']:>8.4f} "
            f"{corr['n']:>5d} "
            f"{strength:>10}"
        )

    print(f"{'='*72}")
    print("  ✓ = p < 0.05 (statistically significant)")
    print("  Strength: |r| > 0.6 = strong, 0.4-0.6 = moderate, < 0.4 = weak")

    # Build markdown report
    lines = [
        "# RAGAS Validation Report",
        "",
        f"**Date:** {datetime.now().strftime('%Y-%m-%d %H:%M')}  ",
        f"**Benchmark:** {bench_path.name}  ",
        f"**Human-labeled rows:** {len(labeled)}",
        "",
        "## Correlation Results",
        "",
        "| Metric | Pearson r | p-value | Spearman r | p-value | n | Strength |",
        "|--------|-----------|---------|------------|---------|---|----------|",
    ]

    for metric, corr in results.items():
        sig = "✓" if (corr["pearson_p"] or 1) < 0.05 else "✗"
        strength = interpret(corr.get("pearson_r"))
        lines.append(
            f"| {metric} | {corr['pearson_r']:.3f} {sig} | {corr['pearson_p']:.4f} | "
            f"{corr['spearman_r']:.3f} | {corr['spearman_p']:.4f} | {corr['n']} | {strength} |"
        )

    lines += [
        "",
        "## Interpretation",
        "",
        "- **|r| > 0.6** (strong): RAGAS metric reliably tracks human quality judgment → use for A/B decisions",
        "- **|r| 0.4–0.6** (moderate): Use as signal but corroborate with human spot-checks",
        "- **|r| < 0.4** (weak): Metric not reliable for this corpus — do not use as primary signal",
        "",
        "## Recommendation",
    ]

    for metric, corr in results.items():
        strength = interpret(corr.get("pearson_r"))
        if strength == "strong":
            lines.append(f"- **{metric}**: trustworthy for ranking experiments on this corpus.")
        elif strength == "moderate":
            lines.append(f"- **{metric}**: use with caution; spot-check low-scoring answers.")
        else:
            lines.append(f"- **{metric}**: not reliable — avoid using as primary eval signal.")

    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d")
    out_path = REPORTS_DIR / f"ragas_validation_{ts}.md"
    out_path.write_text("\n".join(lines), encoding="utf-8")
    print(f"\n  Report saved: {out_path}")


if __name__ == "__main__":
    main()
