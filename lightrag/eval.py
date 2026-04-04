"""LightRAG spike eval — runs 10-question golden sample → RAGAS-scored report.

Usage:
    python tools/lightrag_spike/eval.py
    python tools/lightrag_spike/eval.py --golden tamu_data/evals/golden_sets/golden_20260313_draft_v1_sample10.jsonl

Output:
    tamu_data/evals/reports/lightrag_spike_YYYYMMDD_HHMMSS.md
    tamu_data/evals/reports/lightrag_spike_YYYYMMDD_HHMMSS.xlsx

Requires: ingest.py must have been run first (storage/ must exist).
"""

import argparse
import asyncio
import json
import re
import sys
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Optional

sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))

import config
from lightrag import QueryParam
from rag.tools.langfuse import compute_ragas_metrics

# Add spike dir to path for sibling imports
sys.path.insert(0, str(Path(__file__).resolve().parent))
from wrappers import WORKING_DIR, make_lightrag_improved

_REPO_ROOT = Path(__file__).resolve().parent.parent.parent
REPORTS_DIR = _REPO_ROOT / "tamu_data/evals/reports"
DEFAULT_GOLDEN = _REPO_ROOT / "tamu_data/evals/golden_sets/golden_20260313_draft_v1_sample10.jsonl"


def normalize_course_id(cid: str) -> str:
    """Strip trailing course name: 'CSCE 608 Database Systems' → 'CSCE 608'."""
    match = re.match(r"([A-Z]+ \d+)", cid)
    return match.group(1) if match else cid


@dataclass
class SpikeRow:
    question_id: int
    question: str
    source_course_id: str
    category: str
    stratum: str
    answer: str = ""
    answer_preview: str = ""
    ragas_faithfulness: Optional[float] = None
    ragas_answer_relevancy: Optional[float] = None
    context: str = ""
    elapsed_ms: int = 0
    error: str = ""


def load_golden(path: Path) -> list[dict]:
    rows = []
    with open(path) as f:
        for line in f:
            line = line.strip()
            if line:
                rows.append(json.loads(line))
    return rows


async def run_single(rag, question: str) -> tuple[str, str]:
    """Return (answer, context_str). Two queries: context first, then answer."""
    context = await rag.aquery(question, param=QueryParam(mode="hybrid", only_need_context=True))
    answer = await rag.aquery(question, param=QueryParam(mode="hybrid"))
    return answer, str(context)


async def run_eval(golden_path: Path, storage_dir: Path | None = None, skip_ragas: bool = False) -> list[SpikeRow]:
    wd = storage_dir if storage_dir is not None else WORKING_DIR
    if not wd.exists():
        print(f"ERROR: storage not found at {wd}")
        print("Run ingest.py first.")
        sys.exit(1)

    golden = load_golden(golden_path)
    print(f"Loaded {len(golden)} questions from {golden_path}")

    rag = make_lightrag_improved(working_dir=wd)
    await rag.initialize_storages()
    rows: list[SpikeRow] = []

    for i, entry in enumerate(golden, 1):
        question = entry["question"]
        course_id = normalize_course_id(entry.get("source_course_id", ""))
        category = entry.get("category", "")
        stratum = entry.get("stratum", "")

        print(f"[{i}/{len(golden)}] {question[:80]}...")
        row = SpikeRow(
            question_id=i,
            question=question,
            source_course_id=course_id,
            category=category,
            stratum=stratum,
        )

        t0 = time.time()
        try:
            answer, context = await run_single(rag, question)
            row.answer = answer
            row.context = context
            row.answer_preview = answer[:200].replace("\n", " ")
            row.elapsed_ms = int((time.time() - t0) * 1000)

            if not skip_ragas:
                print(f"  Scoring with RAGAS...")
                scores = compute_ragas_metrics(
                    question=question,
                    contexts=[context] if context else ["(no context)"],
                    answer=answer,
                )
                row.ragas_faithfulness = scores.get("faithfulness")
                row.ragas_answer_relevancy = scores.get("answer_relevancy")
                faith_str = f"{row.ragas_faithfulness:.3f}" if row.ragas_faithfulness is not None else "N/A"
                rel_str = f"{row.ragas_answer_relevancy:.3f}" if row.ragas_answer_relevancy is not None else "N/A"
                print(f"  faithfulness={faith_str}  relevancy={rel_str}")

        except Exception as e:
            row.error = str(e)
            row.elapsed_ms = int((time.time() - t0) * 1000)
            print(f"  ERROR: {e}")

        rows.append(row)

    return rows


def write_markdown(rows: list[SpikeRow], path: Path, golden_path: Path) -> None:
    scored = [r for r in rows if r.ragas_faithfulness is not None]
    mean_faith = sum(r.ragas_faithfulness for r in scored) / len(scored) if scored else float("nan")
    mean_rel = sum(r.ragas_answer_relevancy for r in scored) / len(scored) if scored else float("nan")
    errors = [r for r in rows if r.error]

    lines = [
        "# LightRAG Spike Eval Report",
        "",
        f"**Golden set:** `{golden_path}`  ",
        "**Mode:** hybrid  ",
        f"**LLM:** `{config.TAMU_MODEL}` (TAMU gateway)  ",
        "**Embedder:** voyage-3  ",
        f"**Date:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "",
        "## Summary",
        "",
        "| Metric | Value |",
        "|--------|-------|",
        f"| Questions | {len(rows)} |",
        f"| Scored | {len(scored)} |",
        f"| Errors | {len(errors)} |",
        f"| Mean Faithfulness | {mean_faith:.3f} |",
        f"| Mean Answer Relevancy | {mean_rel:.3f} |",
        "",
        "## Per-Query Results",
        "",
        "| # | Course | Category | Faithfulness | Relevancy | Answer Preview |",
        "|---|--------|----------|-------------|-----------|----------------|",
    ]
    for r in rows:
        faith = f"{r.ragas_faithfulness:.3f}" if r.ragas_faithfulness is not None else "err"
        rel = f"{r.ragas_answer_relevancy:.3f}" if r.ragas_answer_relevancy is not None else "err"
        preview = r.answer_preview[:80].replace("|", "\\|") if r.answer_preview else r.error[:80]
        lines.append(f"| {r.question_id} | {r.source_course_id} | {r.category} | {faith} | {rel} | {preview} |")

    path.write_text("\n".join(lines), encoding="utf-8")
    print(f"Markdown report: {path}")


def write_excel(rows: list[SpikeRow], path: Path, golden_path: Path) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        print("ERROR: openpyxl required. Run: pip install openpyxl")
        return

    wb = openpyxl.Workbook()

    # --- Summary tab ---
    ws_sum = wb.active
    ws_sum.title = "Summary"
    scored = [r for r in rows if r.ragas_faithfulness is not None]
    mean_faith = sum(r.ragas_faithfulness for r in scored) / len(scored) if scored else None
    mean_rel = sum(r.ragas_answer_relevancy for r in scored) / len(scored) if scored else None
    errors = [r for r in rows if r.error]

    summary_data = [
        ("Questions", len(rows)),
        ("Scored", len(scored)),
        ("Errors", len(errors)),
        ("Mean Faithfulness", round(mean_faith, 4) if mean_faith is not None else "N/A"),
        ("Mean Answer Relevancy", round(mean_rel, 4) if mean_rel is not None else "N/A"),
        ("LLM", config.TAMU_MODEL),
        ("Embedder", "voyage-3"),
        ("Mode", "hybrid"),
        ("Golden Set", str(golden_path)),
        ("Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(color="FFFFFF", bold=True)
    ws_sum.append(["Metric", "Value"])
    for cell in ws_sum[1]:
        cell.fill = header_fill
        cell.font = header_font
    for key, val in summary_data:
        ws_sum.append([key, val])
    ws_sum.column_dimensions["A"].width = 30
    ws_sum.column_dimensions["B"].width = 50

    # --- Per-Query tab ---
    ws_pq = wb.create_sheet("Per-Query")
    pq_headers = [
        "q#", "question", "source_course_id", "category", "stratum",
        "answer_preview", "ragas_faithfulness", "ragas_answer_relevancy",
        "elapsed_ms", "error",
    ]
    ws_pq.append(pq_headers)
    for cell in ws_pq[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in rows:
        ws_pq.append([
            r.question_id,
            r.question,
            r.source_course_id,
            r.category,
            r.stratum,
            r.answer_preview,
            r.ragas_faithfulness,
            r.ragas_answer_relevancy,
            r.elapsed_ms,
            r.error,
        ])
    for col in ws_pq.iter_cols(min_row=1, max_row=1):
        ws_pq.column_dimensions[col[0].column_letter].width = 20
    ws_pq.column_dimensions["B"].width = 60
    ws_pq.column_dimensions["F"].width = 80

    # --- Full Answers tab ---
    ws_full = wb.create_sheet("Full Answers")
    ws_full.append(["q#", "question", "source_course_id", "full_answer", "ragas_faithfulness", "ragas_answer_relevancy"])
    for cell in ws_full[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in rows:
        ws_full.append([
            r.question_id,
            r.question,
            r.source_course_id,
            r.answer,
            r.ragas_faithfulness,
            r.ragas_answer_relevancy,
        ])
    ws_full.column_dimensions["D"].width = 120
    for row_cells in ws_full.iter_rows(min_row=2):
        for cell in row_cells:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(path)
    print(f"Excel report: {path}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Run LightRAG eval against golden sample")
    parser.add_argument(
        "--golden",
        type=Path,
        default=DEFAULT_GOLDEN,
        help=f"Path to golden set JSONL (default: {DEFAULT_GOLDEN})",
    )
    parser.add_argument(
        "--storage-dir",
        type=Path,
        default=None,
        help="LightRAG storage dir (default: tools/lightrag_spike/storage/)",
    )
    parser.add_argument(
        "--no-ragas",
        action="store_true",
        help="Skip RAGAS scoring (faster, avoids asyncio conflicts)",
    )
    args = parser.parse_args()

    if not args.golden.exists():
        print(f"ERROR: golden set not found: {args.golden}")
        sys.exit(1)

    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    md_path = REPORTS_DIR / f"lightrag_spike_{timestamp}.md"
    xlsx_path = REPORTS_DIR / f"lightrag_spike_{timestamp}.xlsx"

    rows = asyncio.run(run_eval(args.golden, storage_dir=args.storage_dir, skip_ragas=args.no_ragas))

    write_markdown(rows, md_path, args.golden)
    write_excel(rows, xlsx_path, args.golden)

    # Print summary
    scored = [r for r in rows if r.ragas_faithfulness is not None]
    if scored:
        mean_faith = sum(r.ragas_faithfulness for r in scored) / len(scored)
        mean_rel = sum(r.ragas_answer_relevancy for r in scored) / len(scored)
        print(f"\nMean Faithfulness:      {mean_faith:.3f}")
        print(f"Mean Answer Relevancy:  {mean_rel:.3f}")


if __name__ == "__main__":
    main()
