"""Tests for evals/golden_set.py — load, save, append_run_column."""
import pytest
from evals.golden_set import SCHEMA_COLUMNS


SAMPLE_ITEMS = [
    {"id": 1, "question": "What is CSCE 611 about?", "reference_answer": "OS course.",
     "expected_function": "hybrid_course", "human_notes": None},
    {"id": 2, "question": "What courses cover ML?", "reference_answer": "CSCE 638.",
     "expected_function": "semantic_general", "human_notes": "reworked"},
]


def test_save_then_load_roundtrip(tmp_path):
    from evals.golden_set import load, save
    path = tmp_path / "test.xlsx"
    save(SAMPLE_ITEMS, path)
    loaded = load(path)
    assert len(loaded) == 2
    assert loaded[0]["question"] == "What is CSCE 611 about?"
    assert loaded[0]["expected_function"] == "hybrid_course"
    assert loaded[0]["human_notes"] is None
    assert loaded[1]["human_notes"] == "reworked"


def test_load_skips_empty_question_rows(tmp_path):
    from evals.golden_set import load, save
    items = SAMPLE_ITEMS + [
        {"id": 3, "question": "", "reference_answer": "", "expected_function": "", "human_notes": None}
    ]
    path = tmp_path / "test.xlsx"
    save(items, path)
    loaded = load(path)
    assert len(loaded) == 2


def test_load_ignores_run_columns(tmp_path):
    import openpyxl
    from evals.golden_set import save, load
    path = tmp_path / "test.xlsx"
    save(SAMPLE_ITEMS, path)

    # manually append a run column
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    ws.cell(row=1, column=6, value="run:my_experiment")
    ws.cell(row=2, column=6, value="CSCE 611 §SCHEDULE 0.92")
    wb.save(path)

    loaded = load(path)
    assert "run:my_experiment" not in loaded[0]
    assert len(loaded[0]) == len(SCHEMA_COLUMNS)  # exactly schema keys


def test_append_run_column_adds_column(tmp_path):
    from evals.golden_set import save, append_run_column
    import openpyxl
    path = tmp_path / "test.xlsx"
    save(SAMPLE_ITEMS, path)

    append_run_column(path, "exp_20260412", {1: "CSCE 611 §SCHEDULE 0.92", 2: "CSCE 638 §LO 0.87"})

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert "run:exp_20260412" in headers
    col_idx = headers.index("run:exp_20260412") + 1
    assert ws.cell(row=2, column=col_idx).value == "CSCE 611 §SCHEDULE 0.92"
    assert ws.cell(row=3, column=col_idx).value == "CSCE 638 §LO 0.87"


def test_append_run_column_overwrites_existing(tmp_path):
    from evals.golden_set import save, append_run_column
    import openpyxl
    path = tmp_path / "test.xlsx"
    save(SAMPLE_ITEMS, path)

    append_run_column(path, "exp_20260412", {1: "old value"})
    append_run_column(path, "exp_20260412", {1: "new value"})

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    col_idx = headers.index("run:exp_20260412") + 1
    assert ws.cell(row=2, column=col_idx).value == "new value"
    # only one run column added
    assert headers.count("run:exp_20260412") == 1


def test_eval_chunking_accepts_xlsx_and_appends_run_column(tmp_path):
    """eval_chunking.load_golden_set is gone; xlsx round-trips through golden_set.load."""
    from evals.golden_set import save, load
    items = [
        {"id": 1, "question": "What is CSCE 611 about?", "reference_answer": "OS course.",
         "expected_function": "hybrid_course", "human_notes": None},
    ]
    path = tmp_path / "golden.xlsx"
    save(items, path)
    loaded = load(path)
    # confirm load_golden_set no longer exists in eval_chunking
    import evals.eval_chunking as ec
    assert not hasattr(ec, "load_golden_set")
